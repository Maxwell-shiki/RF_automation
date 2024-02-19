import sys, os
import subprocess
import pyvisa as visa
import openpyxl
import time
import matplotlib.pyplot as plt

# ========= import modules ================================
path_modules = '..\\..\\modules'

for root, dirs, files in os.walk(path_modules):
    for directory in dirs:
        path_module = os.path.join(root, directory)
        sys.path.append(path_module)
        
from DCPowerSupply_ES3631A import DCPowerSupply_ES3631A  
from Multimeter_3458A import Multimeter_3458A

# =========================================================

def main():
    # Start connection
    PS_resource_name = "USB0::0x2A8D::0x1002::MY61003060::0::INSTR"     # ES3631A 直流电源
    DCPS = DCPowerSupply_ES3631A(PS_resource_name)
    MM_resource_name = "GPIB1::22::INSTR"
    MM = Multimeter_3458A(MM_resource_name)

    # 万用表量程设置
    MM.set_measure(set_mode='DCV', set_range=10)
    # 电压列表
    vin_list = []
    vout_list = []

    # 设置恒定电压
    # print()
    # DCPS.set_voltage(2.5, channel = 2)

    # 设置voltage循环变化 1.6V~2.5V, 步长为0.05V
    print()
    start_voltage = 1.2
    end_voltage = 3.7
    step_voltage = 0.01

    # 设置每步变化时的延时时间
    # delay_seconds = 0.1

    current_voltage = start_voltage
    while current_voltage <= end_voltage:
        DCPS.set_voltage(current_voltage, channel = 1)
        print("    Voltage set = ", DCPS.get_voltage(channel = 1), "V")
        vin_list.append(current_voltage)

        voltage = MM.get_data()
        vout_list.append(voltage)
        print("    Voltage output = ", voltage, "V")

        print()
        # time.sleep(delay_seconds)
        current_voltage += step_voltage

    # 画图
    plt.plot(vin_list, vout_list, 'r-')
    plt.xlabel('Vin (V)')
    plt.ylabel('Vout (V)')
    plt.title('Vin-Vout')
    plt.show()

    # 打印数据列表到excel
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Vin-Vout'
    sheet['A1'] = 'Vin (V)'
    sheet['B1'] = 'Vout (V)'
    for i in range(len(vin_list)):
        sheet.cell(row=i+2, column=1).value = vin_list[i]
        sheet.cell(row=i+2, column=2).value = vout_list[i]
    wb.save('./data/case3(1.2~3.7V).xlsx')

    print()
    DCPS.close()
    # MM.close()

    print('  Test done.\n')

if __name__ == "__main__":
    main()
